from __future__ import annotations

import io
import math
import re
import unicodedata
import zipfile
from collections import OrderedDict
from dataclasses import dataclass
from datetime import date, datetime, time
from typing import Callable, Iterable, Mapping, Sequence
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


MCO_ITEM = "Till Compliance"
SCO_ITEM = "Till Compliance - SCO"
REPORT_TIMEZONE = ZoneInfo("Europe/London")
UTF8_BOM = b"\xef\xbb\xbf"
MCO_HIDDEN_COLUMNS: tuple[str, ...] = ("BT", "BV", "BW", "CH")


class ReportGenerationError(ValueError):
    """Raised when the uploaded export cannot produce the required reports."""


@dataclass(frozen=True)
class ColumnSpec:
    question_id: str | None
    header: str
    aliases: tuple[str, ...] = ()
    kind: str = "text"
    required: bool = False
    derived: str | None = None


@dataclass
class GenerationResult:
    files: OrderedDict[str, bytes]
    stats: dict[str, int]
    warnings: list[str]
    reporting_label: str

    def as_zip(self) -> bytes:
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for filename, contents in self.files.items():
                archive.writestr(filename, contents)
        return output.getvalue()


def source_column(
    header: str,
    *aliases: str,
    question_id: str | None = None,
    kind: str = "text",
    required: bool = False,
) -> ColumnSpec:
    return ColumnSpec(
        question_id=question_id,
        header=header,
        aliases=tuple(aliases) if aliases else (header,),
        kind=kind,
        required=required,
    )


def blank_column(header: str, question_id: str | None = None) -> ColumnSpec:
    return ColumnSpec(
        question_id=question_id,
        header=header,
        derived="blank",
    )


MCO_SCHEMA: tuple[ColumnSpec, ...] = (
    source_column("order_internal_id", required=True),
    source_column("order_schedule_type", required=True),
    source_column("client_name", required=True),
    source_column("internal_id", required=True),
    source_column("site_internal_id", required=True),
    source_column("order_end_date", "end_date", kind="date", required=True),
    source_column("site_name", required=True),
    source_column("site_address_1", required=True),
    source_column("site_address_2", required=True),
    source_column("site_address_3", required=True),
    source_column("site_post_code", required=True),
    source_column("submitted_date", kind="date", required=True),
    source_column("approval_date", kind="date", required=True),
    source_column("item_to_order", required=True),
    source_column("date_of_visit", kind="date", required=True),
    source_column("time_of_visit", "time_of_visit_local", "time_of_visit", kind="time", required=True),
    source_column("survey_score", kind="percent", required=True),
    source_column("site_code", kind="number", required=True),
    source_column("primary_result", required=True),
    ColumnSpec(None, "primary_result", kind="text", required=True, derived="result_initial"),
    source_column(
        "Were you able to successfully conduct this audit?",
        question_id="Q1",
    ),
    source_column(
        "Please detail why you were unable to conduct this audit:",
        question_id="Q2",
    ),
    source_column(
        "Please upload your receipt image:",
        "Please upload a photo of your full receipt showing any visible codes:",
        "Please upload your receipt image:",
        question_id="Q3",
    ),
    source_column(
        "Please input the store code from the receipt:",
        question_id="Q2968",
        kind="number",
    ),
    source_column(
        "Please input the transaction code from the receipt:",
        question_id="Q2969",
        kind="number",
    ),
    source_column(
        "Please input the time you entered the store:",
        question_id="Q2852",
        kind="time",
    ),
    source_column("Was there a trolley available?", question_id="Q3073"),
    blank_column("Did you find all the primary items on the shopping list?", "Q2853"),
    blank_column("If 'No', what were you unable to find?", "Q2854"),
    source_column("Select the first item you purchased:", question_id="Q2855"),
    source_column(
        "Please enter the displayed price of the first item:",
        question_id="Q2856",
        kind="number",
    ),
    source_column("Select the second item you purchased:", question_id="Q2857"),
    source_column(
        "Please enter the displayed price of the second item:",
        question_id="Q2858",
        kind="number",
    ),
    source_column("Select the third item you purchased:", question_id="Q2859"),
    source_column(
        "Please enter the displayed price of the third item:",
        question_id="Q2860",
        kind="number",
    ),
    source_column("Select the fourth item you purchased:", question_id="Q2861"),
    source_column(
        "Please enter the displayed price of the fourth item:",
        question_id="Q2862",
        kind="number",
    ),
    source_column(
        "How many manned checkouts does the store have?",
        question_id="Q2876",
        kind="number",
    ),
    source_column(
        "How many manned checkouts were open as you joined the queue?",
        question_id="Q2877",
        kind="number",
    ),
    source_column(
        "How many people were queuing behind the till belt (across all open manned checkouts)?",
        question_id="Q2878",
        kind="number",
    ),
    source_column("Were there any self-scan checkout tills in the store?", question_id="Q2970"),
    source_column(
        "How many manned checkouts had coinage/money roll(s) left visible on any of the checkouts?",
        question_id="Q2881",
        kind="number",
    ),
    source_column(
        "How many closed manned checkouts had their till gates open?",
        question_id="Q2882",
        kind="number",
    ),
    source_column("What number till did you get served at?", question_id="Q2883", kind="number"),
    source_column("What time did you get served?", question_id="Q2884", kind="time"),
    source_column(
        "Was a carrier bag available for your to pick up & leave in your trolley or did you have to ask the cashier for one?",
        question_id="Q2885",
    ),
    source_column("Was the colleague that served you wearing a headset?", question_id="Q2886"),
    source_column("Was the colleague that served you wearing a name badge?", question_id="Q2887"),
    source_column("Was the colleague that served you wearing Lidl uniform?", question_id="Q2888"),
    source_column("Did the colleague wait for the belt to be fully loaded and greet you?", question_id="Q2889"),
    source_column(
        "Did the colleague review the queue e.g. react if another till is required to open?",
        question_id="Q2890",
    ),
    source_column("Did the colleague engage with you? ", "Did the colleague engage with you?", question_id="Q2891"),
    source_column("Did the colleague place the trolley in the 'best' loading position", question_id="Q2892"),
    source_column("What did the cashier do with the items left in the trolley?", question_id="Q2893"),
    source_column("Did the colleague ask you whether you are a Lidl Plus member?", question_id="Q2894"),
    source_column(
        "Did the colleague thank you for shopping with them and say goodbye? ",
        "Did the colleague thank you for shopping with them and say goodbye?",
        question_id="Q2895",
    ),
    source_column("Were you given the correct change?", question_id="Q2896"),
    source_column("If not, what was the discrepancy?", question_id="Q2897", kind="number"),
    source_column(
        "Please detail the exact amount you paid with, the total of the shop, the change you received back and the discrepancy (including whether this was more or less than you should have received):",
        question_id="Q3074",
    ),
    source_column("Were you charged for all items?", question_id="Q2898"),
    source_column("If 'No', which items were you not charged for?", question_id="Q2899"),
    blank_column(
        "Were you charged for all the other items in your basket from the briefing documents list 4?",
        "Q4587",
    ),
    source_column("Input the total value of items not charged for:", question_id="Q2960", kind="number"),
    source_column(
        "Were you charged correctly for all items?",
        "Were you charged for the correct items and the correct quantity of each item in your basket?",
        "Were you charged correctly for all items?",
        question_id="Q2935",
    ),
    source_column(
        "How many items were incorrectly charged?",
        question_id="Q2936",
        kind="number",
    ),
    source_column(
        "Please select the first item that you were incorrectly charged for:",
        question_id="Q2937",
    ),
    source_column(
        "Input the price you were incorrectly charged for the first item:",
        question_id="Q2938",
        kind="number",
    ),
    source_column(
        "Were you overcharged or undercharged for the first item:",
        question_id="Q3076",
    ),
    source_column(
        "Please select the second item that you were incorrectly charged for:",
        question_id="Q2939",
    ),
    source_column(
        "Input the price you were incorrectly charged for the second item:",
        question_id="Q2945",
        kind="number",
    ),
    source_column(
        "Were you overcharged or undercharged for the second item:",
        question_id="Q3077",
    ),
    source_column(
        "Please select the third item that you were incorrectly charged for:",
        question_id="Q2946",
    ),
    source_column(
        "Input the price you were incorrectly charged for the third item:",
        question_id="Q2947",
        kind="number",
    ),
    source_column(
        "Were you overcharged or undercharged for the third item:",
        question_id="Q3078",
    ),
    source_column(
        "Please select the fourth item that you were incorrectly charged for:",
        question_id="Q2948",
    ),
    source_column(
        "Input the price you were incorrectly charged for the fourth item:",
        question_id="Q2949",
        kind="number",
    ),
    source_column(
        "Were you overcharged or undercharged for the fourth item:",
        question_id="Q3079",
    ),
    blank_column("Please select the fifth item that you were incorrectly charged for:", "Q2950"),
    blank_column("Input the price you were incorrectly charged for the fifth item:", "Q2951"),
    blank_column("Were you overcharged or undercharged for the fifth item:", "Q3080"),
    blank_column("Please select the sixth item that you were incorrectly charged for:", "Q2952"),
    blank_column("Input the price you were incorrectly charged for the sixth item:", "Q2953"),
    blank_column("Were you overcharged or undercharged for the sixth item:", "Q3081"),
    blank_column("Please select the seventh item that you were incorrectly charged for:", "Q2954"),
    blank_column("Input the price you were incorrectly charged for the seventh item:", "Q2955"),
    blank_column("Were you overcharged or undercharged for the seventh item:", "Q3082"),
    blank_column(
        "Were you correctly charged for all the items in your basket from the briefing documents list 4 :",
        "Q4588",
    ),
    source_column(
        "Please detail the product, the displayed cost and the charged cost:",
        question_id="Q4589",
    ),
    source_column(
        "Please input the total difference of the items incorrectly charged:",
        question_id="Q2961",
        kind="number",
    ),
    source_column("Please input the total value of receipt:", question_id="Q2958", kind="number"),
    source_column(
        "Please confirm the total amount of VAT you paid on your purchases:",
        question_id="Q3170",
        kind="number",
    ),
    source_column(
        "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
        question_id="Q23",
    ),
)


SCO_SCHEMA: tuple[ColumnSpec, ...] = (
    source_column("order_internal_id", required=True),
    source_column("order_schedule_type", required=True),
    source_column("client_name", required=True),
    source_column("internal_id", required=True),
    # These two headings intentionally reproduce the supplied SCO report.
    source_column("order_end_date", "site_internal_id", required=True),
    source_column("site_internal_id", "end_date", kind="date", required=True),
    source_column("site_name", required=True),
    source_column("site_address_1", required=True),
    source_column("site_address_2", required=True),
    source_column("site_address_3", required=True),
    source_column("site_post_code", required=True),
    source_column("submitted_date", kind="date", required=True),
    source_column("approval_date", kind="date", required=True),
    source_column("item_to_order", required=True),
    source_column("date_of_visit", kind="date", required=True),
    source_column("time_of_visit", "time_of_visit_local", "time_of_visit", kind="time", required=True),
    source_column("site_code", kind="number", required=True),
    source_column("primary_result", required=True),
    source_column(
        "Were you able to successfully conduct this audit?",
        question_id="Q1",
    ),
    source_column(
        "Please detail why you were unable to conduct this audit:",
        question_id="Q2",
    ),
    source_column("Please upload your receipt image:", question_id="Q3"),
    source_column(
        "Please upload a photo of your trolley, before you are at the till, which clearly shows all items you are purchasing:",
        question_id="Q4107",
    ),
    source_column(
        "Please input the store code from the receipt:",
        question_id="Q2968",
        kind="number",
    ),
    source_column(
        "Please input the transaction code from the receipt:",
        question_id="Q2969",
        kind="number",
    ),
    source_column(
        "Please confirm the number on the Self-scan till you used:",
        question_id="Q4439",
        kind="number",
    ),
    source_column(
        "Please input the time you entered the store:",
        question_id="Q2852",
        kind="time",
    ),
    source_column("Inventory", question_id="Q4767"),
    source_column(
        "Was the SCO coordinator present at the time of the audit? ",
        "Was the SCO coordinator present at the time of the audit?",
        question_id="Q4440",
    ),
    source_column(
        "Did the SCO till trigger an intervention (require a cashier to assist)?",
        "When you placed your items on the till, did it trigger an intervention for the cashier to respond to?",
        "Did the SCO till trigger an intervention (require a cashier to assist)?",
        question_id="Q4449",
    ),
    source_column(
        "Did the cashier ensure the second duplicate item was correctly charged for? ",
        "Did the cashier ensure the duplicate item was correctly charged for?",
        "Did the cashier ensure the second duplicate item was correctly charged for?",
        question_id="Q4768",
    ),
    source_column("Please explain what the cashier did:", question_id="Q4769"),
    source_column("SCO Cashier / Customer Service", question_id="Q4770"),
    source_column("Was the cashier who served you wearing a name badge?", question_id="Q68"),
    source_column("Was the cashier wearing a Lidl uniform?", question_id="Q4451"),
    source_column("Was the cashier wearing a high-vis vest?", question_id="Q4452"),
    source_column("Was the cashier wearing a headset?", question_id="Q4453"),
    source_column("Was the cashier wearing a body worn camera?", question_id="Q4454"),
    source_column("Were there carrier bags available at the SCO till?", question_id="Q4455"),
    source_column("How many SCO tills were open?", question_id="Q4456", kind="number"),
    source_column(
        "If there were more than 7 SCO tills open, how many SCO co-ordinators were present? ",
        "If there were more than 7 SCO tills open, how many SCO co-ordinators were present?",
        question_id="Q4457",
        kind="number",
    ),
    source_column("Please input the total value of receipt:", question_id="Q2958", kind="number"),
    source_column(
        "Please confirm the total amount of VAT you paid on your purchases:",
        question_id="Q3170",
        kind="number",
    ),
    source_column("Did you show the letter of authorisation at this site?", question_id="Q3211"),
    source_column("Please detail why you showed the letter of authorisation", question_id="Q3212"),
    source_column(
        "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
        question_id="Q23",
    ),
    blank_column(
        "If the bakery item did not trigger an intervention, your second duplicate items should have triggered an intervention. Did this occur?",
        "Q5579",
    ),
)


INCORRECT_CHARGE_SCHEMA: tuple[ColumnSpec, ...] = (
    source_column("site_name", required=True),
    source_column("site_post_code", required=True),
    source_column("site_code", kind="number", required=True),
    source_column(
        "How many items were incorrectly charged?",
        question_id="Q2936",
        kind="number",
    ),
    source_column(
        "Please select the first item that you were incorrectly charged for:",
        question_id="Q2937",
    ),
    source_column(
        "Input the price you were incorrectly charged for the first item:",
        question_id="Q2938",
        kind="number",
    ),
    source_column(
        "Were you overcharged or undercharged for the first item:",
        question_id="Q3076",
    ),
    source_column(
        "Please select the second item that you were incorrectly charged for:",
        question_id="Q2939",
    ),
    source_column(
        "Input the price you were incorrectly charged for the second item:",
        question_id="Q2945",
        kind="number",
    ),
    source_column(
        "Were you overcharged or undercharged for the second item:",
        question_id="Q3077",
    ),
    source_column(
        "Please select the third item that you were incorrectly charged for:",
        question_id="Q2946",
    ),
    source_column(
        "Input the price you were incorrectly charged for the third item:",
        question_id="Q2947",
        kind="number",
    ),
    source_column(
        "Please input the total difference of the items incorrectly charged:",
        question_id="Q2961",
        kind="number",
        required=True,
    ),
    source_column(
        "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
        question_id="Q23",
    ),
)


ABORT_COLUMNS: tuple[str, ...] = (
    "site_name",
    "site_post_code",
    "site_code",
    "primary_result",
    "Were you able to successfully conduct this audit?",
    "Please detail why you were unable to conduct this audit:",
)


MCO_WIDTHS: tuple[float, ...] = (
    20.28515625, 9.140625, 13.0, 13.0, 13.0, 10.85546875, 27.5703125, 9.140625,
    13.0, 13.0, 13.0, 12.42578125, 13.0, 9.140625, 12.42578125, 9.5703125,
    13.0, 13.0, 9.140625, 13.0, 13.0, 13.0, 9.5703125, 10.28515625,
    9.5703125, 9.140625, 13.0, 13.0, 13.0, 9.5703125, 9.140625, 9.5703125,
    9.140625, 9.5703125, 9.140625, 9.5703125, 13.0, 9.28515625, 13.0,
    9.42578125, 9.42578125, 9.5703125, 13.0, 9.28515625, 13.0, 9.140625,
    13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 70.140625, 9.140625,
    9.42578125, 9.140625, 9.28515625, 9.140625, 9.42578125, 9.140625,
    9.42578125, 64.85546875, 21.140625, 7.7109375, 41.7109375, 10.28515625,
    17.5703125, 47.7109375, 9.42578125, 15.0, 9.140625, 13.0, 11.0,
    9.140625, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0,
    13.0, 9.42578125, 9.42578125, 9.140625, 9.28515625, 13.0, 13.0,
    9.140625,
)

SCO_WIDTHS: tuple[float, ...] = (
    16.140625, 13.0, 13.0, 13.0, 10.5703125, 10.28515625, 13.0, 13.0,
    13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 14.140625, 13.0, 13.0, 13.0,
    13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0,
    13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0,
    13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0,
)

INCORRECT_WIDTHS: tuple[float, ...] = (
    27.5703125, 13.0, 9.28515625, 13.0, 26.42578125, 9.28515625, 13.0,
    13.0, 9.28515625, 13.0, 13.0, 9.28515625, 13.0, 13.0,
)


def _normalise_header(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("&", " and ").replace("’", "'").casefold()
    return " ".join(re.findall(r"[a-z0-9]+", text))


def _column_lookup(df: pd.DataFrame) -> dict[str, list[str]]:
    lookup: dict[str, list[str]] = {}
    for column in df.columns:
        lookup.setdefault(_normalise_header(column), []).append(column)
    return lookup


def _find_column(
    df: pd.DataFrame,
    aliases: Sequence[str],
    lookup: Mapping[str, list[str]] | None = None,
) -> str | None:
    for alias in aliases:
        if alias in df.columns:
            return alias
    lookup = lookup or _column_lookup(df)
    for alias in aliases:
        matches = lookup.get(_normalise_header(alias), [])
        if matches:
            return matches[0]
    return None


def read_export(csv_bytes: bytes) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            df = pd.read_csv(
                io.BytesIO(csv_bytes),
                dtype=str,
                keep_default_na=False,
                encoding=encoding,
                low_memory=False,
            )
            df.columns = [str(column).lstrip("\ufeff").strip() for column in df.columns]
            return df
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ReportGenerationError("The uploaded CSV could not be decoded as UTF-8 or Windows-1252.") from last_error


def _blank(value: object) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value)) or str(value).strip() == ""


def _number(value: object) -> int | float | str | None:
    if _blank(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace(",", "").replace("£", "")
        try:
            number = float(text)
        except ValueError:
            return str(value)
    if not math.isfinite(number):
        return str(value)
    return int(number) if number.is_integer() else number


def _percentage(value: object) -> float | str | None:
    if _blank(value):
        return None
    text = str(value).strip()
    has_percent = text.endswith("%")
    if has_percent:
        text = text[:-1].strip()
    try:
        number = float(text.replace(",", ""))
    except ValueError:
        return str(value)
    if has_percent or number > 1:
        number /= 100
    return number


def _date(value: object) -> datetime | str | None:
    if _blank(value):
        return None
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return str(value)
    if isinstance(parsed, pd.Timestamp):
        return parsed.to_pydatetime().replace(tzinfo=None)
    return parsed


def _time(value: object) -> time | str | None:
    if _blank(value):
        return None
    if isinstance(value, time):
        return value
    parsed = pd.to_datetime(str(value).strip(), errors="coerce")
    if pd.isna(parsed):
        return str(value)
    return parsed.to_pydatetime().time().replace(microsecond=0)


def _convert_series(series: pd.Series, kind: str) -> pd.Series:
    if kind == "number":
        return series.map(_number)
    if kind == "percent":
        return series.map(_percentage)
    if kind == "date":
        return series.map(_date)
    if kind == "time":
        return series.map(_time)
    return series.map(lambda value: None if _blank(value) else value)


def _build_output_frame(
    source: pd.DataFrame,
    schema: Sequence[ColumnSpec],
    report_name: str,
    warnings: list[str],
) -> pd.DataFrame:
    lookup = _column_lookup(source)
    output: OrderedDict[str, pd.Series] = OrderedDict()
    missing_required: list[str] = []

    for index, spec in enumerate(schema):
        output_key = f"{index:03d}:{spec.header}"
        if spec.derived == "blank":
            output[output_key] = pd.Series([None] * len(source), index=source.index, dtype=object)
            continue
        if spec.derived == "result_initial":
            column = _find_column(source, ("primary_result",), lookup)
            if column is None:
                missing_required.append("primary_result")
                output[output_key] = pd.Series([None] * len(source), index=source.index, dtype=object)
            else:
                output[output_key] = source[column].map(
                    lambda value: str(value).strip().lower()[:1] if not _blank(value) else None
                )
            continue

        column = _find_column(source, spec.aliases, lookup)
        if column is None:
            if spec.required:
                missing_required.append(spec.header)
            else:
                warnings.append(f"{report_name}: no source column matched '{spec.header}'; it was left blank.")
            output[output_key] = pd.Series([None] * len(source), index=source.index, dtype=object)
            continue
        output[output_key] = _convert_series(source[column], spec.kind)

    if missing_required:
        missing_text = ", ".join(sorted(set(missing_required)))
        raise ReportGenerationError(f"{report_name} is missing required source columns: {missing_text}")

    return pd.DataFrame(output, index=source.index)


def _stable_sort(
    frame: pd.DataFrame,
    aliases: Sequence[str],
    *,
    numeric: bool,
    warnings: list[str],
    label: str,
) -> pd.DataFrame:
    column = _find_column(frame, aliases)
    if column is None:
        warnings.append(f"The {label} sort was skipped because its source column was not found.")
        return frame
    if numeric:
        key = pd.to_numeric(frame[column].astype(str).str.strip(), errors="coerce")
    else:
        key = frame[column].map(lambda value: pd.NA if _blank(value) else str(value).strip().casefold())
    temporary = frame.assign(__sort_key=key)
    return temporary.sort_values("__sort_key", kind="mergesort", na_position="last").drop(columns="__sort_key")


def _write_workbook(
    frame: pd.DataFrame,
    schema: Sequence[ColumnSpec],
    *,
    sheet_name: str,
    widths: Sequence[float],
    font_name: str,
    zoom: int | None,
    hidden_columns: Sequence[str] = (),
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    if zoom is not None:
        worksheet.sheet_view.zoomScale = zoom

    font = Font(name=font_name, size=11)
    for column_index, spec in enumerate(schema, start=1):
        if spec.question_id:
            worksheet.cell(1, column_index, spec.question_id)
        worksheet.cell(2, column_index, spec.header)
        worksheet.cell(1, column_index).font = font
        worksheet.cell(2, column_index).font = font
        if column_index <= len(widths):
            worksheet.column_dimensions[get_column_letter(column_index)].width = widths[column_index - 1]

    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=3):
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column_index, value)
            cell.font = font
            kind = schema[column_index - 1].kind
            if kind == "date" and isinstance(value, (date, datetime)):
                cell.number_format = "mm-dd-yy"
            elif kind == "time" and isinstance(value, time):
                cell.number_format = "h:mm"
            elif kind == "percent" and isinstance(value, (int, float)):
                cell.number_format = "0.00%"

    last_column = get_column_letter(len(schema))
    last_row = max(2, len(frame) + 2)
    worksheet.auto_filter.ref = f"A2:{last_column}{last_row}"
    for column in hidden_columns:
        worksheet.column_dimensions[column].hidden = True
    workbook.active = 0

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _abort_csv(frame: pd.DataFrame) -> bytes:
    missing = [column for column in ABORT_COLUMNS if column not in frame.columns]
    if missing:
        raise ReportGenerationError(
            "Abort reports are missing required source columns: " + ", ".join(missing)
        )
    text = frame.loc[:, ABORT_COLUMNS].to_csv(index=False, lineterminator="\r\n")
    return text.encode("utf-8-sig")


def encrypt_xlsx(xlsx_bytes: bytes, password: str) -> bytes:
    if not password:
        raise ReportGenerationError("The MD password is empty.")
    try:
        from msoffcrypto.format.ooxml import OOXMLFile
    except ImportError as exc:
        raise ReportGenerationError(
            "msoffcrypto-tool is not installed, so the MD workbooks cannot be encrypted."
        ) from exc

    plain = io.BytesIO(xlsx_bytes)
    encrypted = io.BytesIO()
    OOXMLFile(plain).encrypt(password, encrypted)
    return encrypted.getvalue()


def generate_reports(
    source: pd.DataFrame,
    md_password: str,
    *,
    now: datetime | None = None,
    encryptor: Callable[[bytes, str], bytes] | None = None,
) -> GenerationResult:
    warnings: list[str] = []
    required_core = ("item_to_order", "primary_result")
    missing_core = [column for column in required_core if column not in source.columns]
    if missing_core:
        raise ReportGenerationError(
            "The export is missing required columns: " + ", ".join(missing_core)
        )

    item_key = source["item_to_order"].map(lambda value: str(value).strip().casefold())
    mco_source = source[item_key == MCO_ITEM.casefold()].copy()
    sco_source = source[item_key == SCO_ITEM.casefold()].copy()
    ignored_count = len(source) - len(mco_source) - len(sco_source)
    if ignored_count:
        warnings.append(
            f"{ignored_count} row(s) were ignored because item_to_order was neither '{MCO_ITEM}' nor '{SCO_ITEM}'."
        )
    if mco_source.empty:
        raise ReportGenerationError(f"No '{MCO_ITEM}' rows were found in the export.")
    if sco_source.empty:
        raise ReportGenerationError(f"No '{SCO_ITEM}' rows were found in the export.")

    mco_sorted = _stable_sort(
        mco_source,
        ("primary_result",),
        numeric=False,
        warnings=warnings,
        label="MCO primary_result",
    )
    sco_sorted = _stable_sort(
        sco_source,
        ("primary_result",),
        numeric=False,
        warnings=warnings,
        label="SCO primary_result",
    )

    mco_is_abort = mco_sorted["primary_result"].astype(str).str.strip().str.casefold().eq("abort")
    sco_is_abort = sco_sorted["primary_result"].astype(str).str.strip().str.casefold().eq("abort")
    mco_aborts = mco_sorted[mco_is_abort].copy()
    sco_aborts = sco_sorted[sco_is_abort].copy()
    mco_report_source = mco_sorted[~mco_is_abort].copy()
    sco_report_source = sco_sorted[~sco_is_abort].copy()

    # These are separate stable sorts, matching the order of operations in the supplied report.
    mco_report_source = _stable_sort(
        mco_report_source,
        ("How many items were incorrectly charged?",),
        numeric=True,
        warnings=warnings,
        label="MCO incorrectly charged item count",
    )
    mco_report_source = _stable_sort(
        mco_report_source,
        ("Please select the first item that you were incorrectly charged for:",),
        numeric=False,
        warnings=warnings,
        label="MCO first incorrectly charged item",
    )
    difference_aliases = ("Please input the total difference of the items incorrectly charged:",)
    mco_report_source = _stable_sort(
        mco_report_source,
        difference_aliases,
        numeric=True,
        warnings=warnings,
        label="MCO total charge difference",
    )

    difference_column = _find_column(mco_report_source, difference_aliases)
    if difference_column is None:
        raise ReportGenerationError(
            "The export does not contain 'Please input the total difference of the items incorrectly charged:'."
        )
    difference_values = pd.to_numeric(
        mco_report_source[difference_column].astype(str).str.strip(), errors="coerce"
    )
    zero_discrepancy_source = mco_report_source[difference_values.eq(0)].copy()

    mco_frame = _build_output_frame(mco_report_source, MCO_SCHEMA, "MCO report", warnings)
    sco_frame = _build_output_frame(sco_report_source, SCO_SCHEMA, "SCO report", warnings)
    incorrect_frame = _build_output_frame(
        zero_discrepancy_source,
        INCORRECT_CHARGE_SCHEMA,
        "Incorrect charge discrepancy report",
        warnings,
    )

    mco_xlsx = _write_workbook(
        mco_frame,
        MCO_SCHEMA,
        sheet_name="Sheet1",
        widths=MCO_WIDTHS,
        font_name="Calibri",
        zoom=85,
        hidden_columns=MCO_HIDDEN_COLUMNS,
    )
    sco_xlsx = _write_workbook(
        sco_frame,
        SCO_SCHEMA,
        sheet_name="audits_basic_data_export (18)",
        widths=SCO_WIDTHS,
        font_name="Aptos Narrow",
        zoom=85,
    )
    incorrect_xlsx = _write_workbook(
        incorrect_frame,
        INCORRECT_CHARGE_SCHEMA,
        sheet_name="Sheet1",
        widths=INCORRECT_WIDTHS,
        font_name="Calibri",
        zoom=None,
    )

    encryptor = encryptor or encrypt_xlsx
    mco_md = encryptor(mco_xlsx, md_password)
    sco_md = encryptor(sco_xlsx, md_password)

    current = now or datetime.now(REPORT_TIMEZONE)
    month_name = current.strftime("%B")
    report_period = current.strftime("%B %y")
    files: OrderedDict[str, bytes] = OrderedDict(
        (
            (f"Lidl MS dataset_wQuestions - {report_period}.xlsx", mco_xlsx),
            (f"Lidl MS dataset_wQuestions - {report_period} MD.xlsx", mco_md),
            (f"Lidl MS SCO dataset_wQuestions - {report_period}.xlsx", sco_xlsx),
            (f"Lidl MS SCO dataset_wQuestions - {report_period} MD.xlsx", sco_md),
            (f"{month_name} - MCO - Aborts.csv", _abort_csv(mco_aborts)),
            (f"{month_name} - SCO - Aborts.csv", _abort_csv(sco_aborts)),
            (f"{month_name}_IncorrectCharge0Discrepency.xlsx", incorrect_xlsx),
        )
    )

    stats = {
        "uploaded_rows": len(source),
        "mco_report_rows": len(mco_report_source),
        "mco_abort_rows": len(mco_aborts),
        "zero_discrepancy_rows": len(zero_discrepancy_source),
        "sco_report_rows": len(sco_report_source),
        "sco_abort_rows": len(sco_aborts),
        "ignored_rows": ignored_count,
    }
    return GenerationResult(
        files=files,
        stats=stats,
        warnings=list(dict.fromkeys(warnings)),
        reporting_label=report_period,
    )


def generate_reports_from_csv(
    csv_bytes: bytes,
    md_password: str,
    *,
    now: datetime | None = None,
    encryptor: Callable[[bytes, str], bytes] | None = None,
) -> GenerationResult:
    return generate_reports(
        read_export(csv_bytes),
        md_password,
        now=now,
        encryptor=encryptor,
    )
